from datetime import datetime
from pathlib import PurePath

import openpyxl

from restaurant_menu_app.tasks.tasks_app import celery_app


@celery_app.task(track_started=True)
def save_data_to_file(data: dict):
    file_name = f"{datetime.now().strftime('%d-%m-%Y-%H-%M')}_restaurant_menu.xlsx"
    file_path = str(PurePath(__file__).parent.parent.parent.joinpath("data", file_name))
    save_to_xlsx(data, file_path)
    return {"file_name": file_name}


def get_result(task_id: str):
    task = celery_app.AsyncResult(task_id)
    return task


def save_to_xlsx(data: dict, file_name: str):
    menus = data
    book = openpyxl.Workbook()
    sheet = book.active
    menu_row, menu_column, menu_counter = 1, 1, 1
    for menu in menus:
        sheet.cell(menu_row, menu_column).value = menu_counter
        sheet.cell(menu_row, menu_column + 1).value = menu["menu_title"]
        sheet.cell(menu_row, menu_column + 2).value = menu["menu_description"]
        menu_counter += 1
        submenu_row = menu_row + 1
        submenu_column = menu_column + 1
        submenu_counter = 1
        if menu["child_submenus"]:
            for submenu in menu["child_submenus"]:
                sheet.cell(submenu_row, submenu_column).value = submenu_counter
                sheet.cell(submenu_row, submenu_column + 1).value = submenu["submenu_title"]
                sheet.cell(submenu_row, submenu_column + 2).value = submenu["submenu_description"]
                submenu_counter += 1
                dish_row = submenu_row + 1
                dish_column = submenu_column + 1
                dish_counter = 1
                if submenu["child_dishes"]:
                    for dish in submenu["child_dishes"]:
                        sheet.cell(dish_row, dish_column).value = dish_counter
                        sheet.cell(dish_row, dish_column + 1).value = dish["dish_title"]
                        sheet.cell(dish_row, dish_column + 2).value = dish["dish_description"]
                        sheet.cell(dish_row, dish_column + 3).value = dish["dish_price"]
                        dish_counter += 1
                        dish_row += 1
                submenu_row = dish_row
        menu_row = submenu_row
    book.save(file_name)
    book.close()
